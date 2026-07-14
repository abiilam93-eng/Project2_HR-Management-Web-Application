from datetime import datetime

from openpyxl import load_workbook
import openpyxl

def get_row_count(file, sheet_name):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def read_data(file, sheet_name, row_num, column_num):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=column_num).value


def write_result(file, sheet_name, row_num, result):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=4).value = datetime.now().strftime("%Y-%m-%d")
    sheet.cell(row=row_num, column=8).value = result
    workbook.save(file)
    workbook.close()

