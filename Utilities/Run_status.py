import openpyxl

def update_scenario_status(file_path, sheet_name, scenario_id, status):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == scenario_id:
            sheet.cell(row=row, column=6).value = status
            break

    wb.save(file_path)